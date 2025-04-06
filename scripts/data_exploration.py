import os
import pandas as pd
import numpy as np
from dotenv import load_dotenv
import re
import matplotlib.pyplot as plt
import seaborn as sns
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from PIL import Image
from sqlalchemy import create_engine
import zipfile
from datetime import datetime

# === Load DB Credentials from .env ===
load_dotenv()
DB_HOST = os.getenv("DB_HOST")
DB_NAME = os.getenv("DB_NAME")
DB_USER = os.getenv("DB_USER")
DB_PASSWORD = os.getenv("DB_PASSWORD")
DB_PORT = os.getenv("DB_PORT")

def get_engine():
    """Create and return a SQLAlchemy engine for database connections"""
    connection_string = f"postgresql://{DB_USER}:{DB_PASSWORD}@{DB_HOST}:{DB_PORT}/{DB_NAME}"
    return create_engine(connection_string)

def unit_conversion_factor(unit):
    if unit is None:
        return None, None
    unit = str(unit).strip().lower()
    replacements = {
        "äq": "eq", "äquiv": "eq", "äqv": "eq", "äqu": "eq", "eqv": "eq", "äquiv.": "eq", "äq.": "eq", "ä": "a",
        " ": "", "v.": "v", "eq.": "eq", "eqv.": "eq", "equiv.": "eq", "_": "", "−": "-", "–": "-"
    }
    for old, new in replacements.items():
        unit = unit.replace(old, new)
    unit = unit.replace("³", "^3").replace("²", "^2")
    patterns = {
        r"^kg.*": ("kg", 1), r"^g.*": ("kg", 1e-3), r"^mg.*": ("kg", 1e-6), r"^t.*": ("kg", 1e3),
        r"^mol.*": ("mol", 1), r"^mole.*": ("mol", 1),
        r"^mj.*": ("MJ", 1), r"^kwh.*": ("MJ", 3.6),
        r"^m3.*": ("m3", 1), r"^m\^3.*": ("m3", 1), r"^l.*": ("m3", 1e-3), r"^cm3.*": ("m3", 1e-6),
        r"^m2.*": ("m2", 1), r"^m\^2.*": ("m2", 1),
        r"^kbq.*": ("kBq", 1), r"^ctue.*": ("CTUe", 1), r"^ctuh.*": ("CTUh", 1), r"^sqp.*": ("SQP", 1),
        r"^diseaseincidence.*": ("disease_incidence", 1), r"^krankheitsfälle.*": ("disease_incidence", 1),
        r"^-+$": (None, None), r"^dimensionless.*": ("dimensionless", 1)
    }
    for pattern, (common, factor) in patterns.items():
        if re.match(pattern, unit):
            return common, factor
    print(f"⚠️ Unrecognized unit: '{unit}'")
    return None, None

def fetch_lcia_module_amounts():
    query = """
        SELECT lr.method_en, lr.indicator_key, lr.unit, lma.module, lma.amount,
               lr.process_id, p.name_en as product_name
        FROM lcia_results lr
        JOIN lcia_moduleamounts lma ON lr.lcia_id = lma.lcia_id
        LEFT JOIN products p ON lr.process_id = p.process_id
        WHERE lma.amount IS NOT NULL
    """
    engine = get_engine()
    return pd.read_sql(query, engine)

def fetch_exchange_module_amounts():
    query = """
        SELECT e.flow_en, e.indicator_key, e.direction, e.unit, ema.module, ema.amount,
               e.process_id, p.name_en as product_name
        FROM exchanges e
        JOIN exchange_moduleamounts ema ON e.exchange_id = ema.exchange_id
        LEFT JOIN products p ON e.process_id = p.process_id
        WHERE ema.amount IS NOT NULL
    """
    engine = get_engine()
    return pd.read_sql(query, engine)

def normalize_units_with_metadata(df, group_cols, value_col="amount"):
    """
    Normalize units by choosing the most common unit for each indicator group.
    
    Parameters:
    df (pandas.DataFrame): Input dataframe containing unit and amount columns
    group_cols (list): Columns to group by, should include the indicator column
    value_col (str): Column name containing the values to normalize
    
    Returns:
    pandas.DataFrame: Dataframe with added columns for original_unit, 
                     common_unit (most frequent), and normalized_amount
    """
    df = df.copy()
    df["original_unit"] = df["unit"]
    df["common_unit"] = None
    df["normalized_amount"] = np.nan
    
    # Process each indicator group separately
    for keys, group in df.groupby(group_cols):
        # Get frequency of each unit in this indicator group
        unit_counts = group["unit"].value_counts().dropna()
        if len(unit_counts) == 0:
            continue
            
        # Select the most frequent unit as the base unit
        base_unit = unit_counts.index[0]
        print(f"For indicator {keys}: Using most common unit '{base_unit}' ({unit_counts[base_unit]} occurrences)")
        
        # Get standardized unit name and conversion factor
        common_unit, base_factor = unit_conversion_factor(base_unit)
        if common_unit is None or base_factor is None:
            print(f"  ⚠️ Could not determine common unit for base unit '{base_unit}'")
            continue
            
        # Normalize all values in this indicator group to the common unit
        for idx, row in group.iterrows():
            unit = row["unit"]
            amount = row[value_col]
            if pd.isna(amount):
                continue
                
            _, factor = unit_conversion_factor(unit)
            if factor is not None:
                # Convert using ratio of conversion factors
                norm_value = amount * (factor / base_factor)
                df.at[idx, "normalized_amount"] = norm_value
                df.at[idx, "common_unit"] = common_unit
            else:
                print(f"  ⚠️ Could not determine conversion factor for unit '{unit}'")
                
    return df

def prepare_lcia(df):
    df = df.copy()
    df["name"] = df["method_en"]
    df["source"] = "LCIA"
    return normalize_units_with_metadata(df, ["indicator_key"])

def prepare_exchanges(df):
    df = df.copy()
    df["name"] = df["flow_en"]
    df["source"] = "Exchange"
    return normalize_units_with_metadata(df, ["indicator_key"])

def summarize_by_module(df):
    # First, find the most common name for each indicator_key
    indicator_names = df.groupby("indicator_key")["name"].agg(
        lambda x: x.value_counts().index[0]
    ).reset_index()
    indicator_names.columns = ["indicator_key", "most_common_name"]
    
    # Group by indicator_key and module for summaries
    module_summary = df.groupby(["indicator_key", "module"]).agg(
        count=("normalized_amount", "count"),
        original_unit=("original_unit", lambda x: ", ".join(sorted(set(x.dropna())))),
        common_unit=("common_unit", "first"),
        mean=("normalized_amount", "mean"),
        std=("normalized_amount", "std"),
        min=("normalized_amount", "min"),
        max=("normalized_amount", "max"),
        median=("normalized_amount", "median")
    ).reset_index()
    
    # Merge to get the most common name for each indicator_key
    module_summary = pd.merge(module_summary, indicator_names, on="indicator_key")
    
    # Rename columns for output
    module_summary["indicator"] = module_summary["indicator_key"]
    module_summary["name"] = module_summary["most_common_name"]
    
    return module_summary[["indicator", "name", "module", "count", "original_unit", "common_unit", "mean", "std", "min", "max", "median"]]

def detect_outliers(df):
    outliers_result = []
    for (key, mod), subdf in df.groupby(["indicator_key", "module"]):
        if len(subdf) < 5:
            continue
        q1 = subdf["normalized_amount"].quantile(0.25)
        q3 = subdf["normalized_amount"].quantile(0.75)
        iqr = q3 - q1
        low = q1 - 1.5 * iqr
        high = q3 + 1.5 * iqr
        mean = subdf["normalized_amount"].mean()
        range_size = high - low
        
        out = subdf[(subdf["normalized_amount"] < low) | (subdf["normalized_amount"] > high)]
        for _, row in out.iterrows():
            # Calculate percentage position in the valid range and deviation from mean
            amount = row["normalized_amount"]
            percent_in_range = None
            
            # For values below the lower bound
            if amount < low:
                # Negative percentage indicates how much below the lower bound
                percent_in_range = -100 * (low - amount) / range_size
            # For values above the upper bound
            elif amount > high:
                # Percentage above 100 indicates how much above the upper bound
                percent_in_range = 100 + (100 * (amount - high) / range_size)
            
            # Calculate deviation from mean (both absolute and percentage)
            abs_deviation = amount - mean
            pct_deviation = (amount - mean) / mean * 100 if mean != 0 else float('inf')
            
            # Get product name or use process ID if not available
            process_id = row["process_id"]
            product_name = row.get("product_name")
            if pd.isna(product_name) or product_name is None or product_name == "":
                product_name = f"Process {process_id}"
            
            outliers_result.append({
                "process_id": process_id,
                "product_name": product_name,
                "indicator": row["indicator_key"],
                "module": row["module"],
                "unit": row["common_unit"],
                "min": low,
                "max": high,
                "mean": mean,
                "amount": row["normalized_amount"],
                "percent_in_range": percent_in_range,
                "abs_deviation": abs_deviation,
                "pct_deviation": pct_deviation,
                "comment": "Outlier by IQR"
            })
    return pd.DataFrame(outliers_result)

def sanitize_filename(name):
    """Sanitize a string to be used as a filename"""
    # Replace invalid characters with underscore
    invalid_chars = r'<>:"/\\|?*'
    for char in invalid_chars:
        name = name.replace(char, '_')
    # Remove any other non-printable or control characters
    name = ''.join(c for c in name if c.isprintable() and not c.isspace())
    # Limit length and ensure not empty
    return name[:50] if name else 'unnamed'

def create_filtered_plots(data, indicator, path_box, path_hist):
    clean_data = data.copy()
    
    # Skip plotting if less than 5 data points
    if len(clean_data) < 5:
        # Create empty plot files to avoid file not found errors
        plt.figure(figsize=(6, 4))
        plt.text(0.5, 0.5, f"Insufficient data for {indicator}", 
                 horizontalalignment='center', verticalalignment='center')
        plt.savefig(path_box)
        plt.savefig(path_hist)
        plt.close()
        return
        
    q1 = clean_data["normalized_amount"].quantile(0.25)
    q3 = clean_data["normalized_amount"].quantile(0.75)
    iqr = q3 - q1
    lower_bound = q1 - 1.5 * iqr
    upper_bound = q3 + 1.5 * iqr
    clean_data = clean_data[(clean_data["normalized_amount"] >= lower_bound) &
                            (clean_data["normalized_amount"] <= upper_bound)]

    plt.figure(figsize=(6, 4))
    sns.boxplot(x='module', y='normalized_amount', data=clean_data)
    plt.title(f"Boxplot: {indicator}")
    plt.tight_layout()
    plt.savefig(path_box)
    plt.close()

    plt.figure(figsize=(6, 4))
    sns.histplot(clean_data["normalized_amount"], bins=30, kde=True)
    plt.title(f"Histogram: {indicator}")
    plt.tight_layout()
    plt.savefig(path_hist)
    plt.close()

def export_combined_excel(combined, outliers):
    # Create directory for plots
    os.makedirs("temp_plots", exist_ok=True)
    
    # Create a new workbook with openpyxl directly
    wb = Workbook()
    # Keep the default sheet for now (we'll use it if needed)
    default_sheet = wb.active
    default_sheet.title = "Summary"
    
    # Create a list to track all files to be zipped
    files_to_zip = []
    excel_filename = "lcia_exchange_analysis.xlsx"
    files_to_zip.append(excel_filename)
    
    # Process each indicator_key/source group
    has_data = False
    
    # Check if we're using the old column names or new ones in outliers DataFrame
    # This ensures backward compatibility
    using_old_columns = 'min' in outliers.columns and 'max' in outliers.columns
    
    # Define column mappings for compatibility
    old_to_new = {
        'min': 'lower_bound',
        'max': 'upper_bound',
        'mean': 'normal_mean'
    }
    new_to_old = {v: k for k, v in old_to_new.items()}
    
    for (indicator_key, source), group in combined.groupby(["indicator_key", "source"]):
        if group.empty:
            continue
            
        has_data = True
        
        # Get most common name for this indicator_key
        most_common_name = group["name"].value_counts().index[0] if not group["name"].empty else indicator_key
        
        # Create sheet name using indicator_key (and source)
        sheet_name = f"{indicator_key} ({source})"
        if len(sheet_name) > 31:  # Excel sheet name length limit
            sheet_name = f"{indicator_key[:28]} ({source})"
            
        sheet = wb.create_sheet(sheet_name)
        
        # Calculate summary by module
        mod_summary = summarize_by_module(group)
        
        # Write summary to worksheet
        rows = dataframe_to_rows(mod_summary, index=False, header=True)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, val in enumerate(row, 1):
                sheet.cell(row=r_idx, column=c_idx).value = val
        
        # Generate plots without outliers - use sanitized filenames with indicator_key
        safe_indicator_key = sanitize_filename(indicator_key)
        path_box = f"temp_plots/{safe_indicator_key}_{source}_boxplot.png"
        path_hist = f"temp_plots/{safe_indicator_key}_{source}_histogram.png"
        create_filtered_plots(group, indicator_key, path_box, path_hist)
        
        # Add the plot files to our zip list
        files_to_zip.append(path_box)
        files_to_zip.append(path_hist)
        
        # Add plots below the summary table with enough space
        summary_rows = len(mod_summary) + 3  # Add some extra space
        
        # Add boxplot
        try:
            img_box = XLImage(path_box)
            img_box.anchor = f"A{summary_rows}"
            sheet.add_image(img_box)
            
            # Add histogram next to boxplot (not on top of any text)
            img_hist = XLImage(path_hist)
            img_hist.anchor = f"I{summary_rows}"
            sheet.add_image(img_hist)
        except Exception as e:
            # If image insertion fails, add a note
            sheet.cell(row=summary_rows, column=1).value = f"Note: Could not insert plots. Error: {str(e)}"
        
        # Add outliers section after the plots - filter by indicator_key
        out = outliers[outliers["indicator"] == indicator_key].sort_values("process_id")
        if not out.empty:
            # Determine where to start the outliers section (below the plots)
            outlier_start_row = summary_rows + 18  # Add enough space for plots
            
            # Add header for outliers section
            sheet.cell(row=outlier_start_row, column=1).value = "OUTLIERS"
            from openpyxl.styles import Font
            sheet.cell(row=outlier_start_row, column=1).font = Font(bold=True)
            
            # Add headers for outliers table with appropriate column names
            if using_old_columns:
                headers = ["process_id", "product_name", "indicator", "module", "unit", 
                         "min", "max", "mean", "amount", "percent_in_range", "pct_deviation", 
                         "normal_process_id", "normal_product_name", "normal_amount", "comment"]
            else:
                headers = ["process_id", "product_name", "indicator", "module", "unit", 
                         "lower_bound", "upper_bound", "normal_min", "normal_max", "normal_mean", "full_dataset_mean", 
                         "amount", "percent_in_range", "pct_deviation", 
                         "normal_process_id", "normal_product_name", "normal_amount", "comment"]
            
            for i, h in enumerate(headers):
                sheet.cell(row=outlier_start_row + 1, column=i + 1).value = h
                sheet.cell(row=outlier_start_row + 1, column=i + 1).font = Font(bold=True)
            
            # Write outliers with empty row when process_id changes
            row_idx = outlier_start_row + 1
            prev_pid = None
            
            for i, row_data in out.iterrows():
                # Insert empty row when process_id changes (except for the first process)
                if row_data['process_id'] != prev_pid and prev_pid is not None:
                    row_idx += 1
                
                row_idx += 1
                prev_pid = row_data['process_id']
                
                # Write outlier data, handling column name differences
                for j, h in enumerate(headers):
                    # Handle column name differences between old and new format
                    if h in row_data:
                        value = row_data[h]
                    elif using_old_columns and h in new_to_old:
                        value = row_data[new_to_old[h]]
                    elif not using_old_columns and h in old_to_new:
                        value = row_data[old_to_new[h]]
                    else:
                        value = None  # For columns that don't exist in current data
                        
                    sheet.cell(row=row_idx, column=j + 1).value = value
    
    # Create "All Outliers" sheet
    if not outliers.empty:
        # Add a sheet with all outliers sorted by frequency
        all_outliers_sheet = wb.create_sheet("All Outliers")
        
        # Count frequency of each process_id in outliers
        process_counts = outliers['process_id'].value_counts().reset_index()
        process_counts.columns = ['process_id', 'occurrence_count']
        
        # Merge counts with outliers dataframe
        outliers_with_counts = pd.merge(outliers, process_counts, on='process_id')
        
        # Sort by occurrence count (descending), then by process_id
        sorted_outliers = outliers_with_counts.sort_values(
            ['occurrence_count', 'process_id'], 
            ascending=[False, True]
        )
        
        # Define appropriate headers based on columns available
        if using_old_columns:
            all_headers = ["process_id", "product_name", "indicator", "module", "unit", 
                          "min", "max", "mean", "amount", "percent_in_range", "pct_deviation", 
                          "normal_process_id", "normal_product_name", "normal_amount", 
                          "occurrence_count", "comment"]
        else:
            all_headers = ["process_id", "product_name", "indicator", "module", "unit", 
                          "lower_bound", "upper_bound", "normal_min", "normal_max", "normal_mean", "full_dataset_mean",
                          "amount", "percent_in_range", "pct_deviation", 
                          "normal_process_id", "normal_product_name", "normal_amount", 
                          "occurrence_count", "comment"]
        
        # Add header row
        for i, h in enumerate(all_headers):
            all_outliers_sheet.cell(row=1, column=i + 1).value = h
            all_outliers_sheet.cell(row=1, column=i + 1).font = Font(bold=True)
        
        # Filter columns to only include those that exist in the dataframe
        existing_headers = [h for h in all_headers if h in sorted_outliers.columns]
        
        # Write all outliers sorted by frequency
        for r_idx, row_data in enumerate(sorted_outliers[existing_headers].itertuples(index=False), 2):
            for c_idx, val in enumerate(row_data, 1):
                all_outliers_sheet.cell(row=r_idx, column=c_idx).value = val

    # If no data was processed, add a message to the default sheet
    if not has_data:
        default_sheet.cell(row=1, column=1).value = "No data found to analyze"
    else:
        # If we have data, we can safely remove the default sheet
        if "Summary" in wb.sheetnames and len(wb.sheetnames) > 1:
            wb.remove(wb["Summary"])
    
    # Save the workbook
    wb.save(excel_filename)
    
    # Create zip file with timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    zip_filename = f"lcia_exchange_analysis_{timestamp}.zip"
    
    with zipfile.ZipFile(zip_filename, 'w') as zipf:
        # Add Excel file
        zipf.write(excel_filename)
        
        # Add all plot files
        for file in files_to_zip:
            if os.path.exists(file):
                zipf.write(file)
    
    print(f"✅ Analysis saved to '{excel_filename}'")
    print(f"✅ All files zipped to '{zip_filename}'")
    
    # Remove temporary plot files
    for f in os.listdir("temp_plots"):
        os.remove(os.path.join("temp_plots", f))
    os.rmdir("temp_plots")

def main():
    lcia_df = prepare_lcia(fetch_lcia_module_amounts())
    exch_df = prepare_exchanges(fetch_exchange_module_amounts())
    combined = pd.concat([lcia_df, exch_df], ignore_index=True)
    outliers = detect_outliers(combined)
    export_combined_excel(combined, outliers)
    print("✅ Analysis saved to 'lcia_exchange_analysis.xlsx'")

if __name__ == "__main__":
    main()