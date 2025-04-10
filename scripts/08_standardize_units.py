import pandas as pd
import psycopg2
import os
import re
from dotenv import load_dotenv
import traceback
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Border, Side

# Load environment variables from .env file (for database credentials)
load_dotenv()

not_found_units = []
def normalize_unit(unit):
    """
    Standardize unit names to a consistent format.
    Comprehensive version that handles all units found in the CSV files.
    """
    import re
    
    if not isinstance(unit, str):
        return unit
    
    # Handle NULL values
    if unit.upper() == "NULL" or unit.strip() == "":
        return "-"
    
    unit = unit.strip().lower()
    
    # Standard mappings with expanded patterns
    mappings = {
        # Volume units
        r"^m[\^³]?[3]?$": "m3",
        r"^m\s*[3³]$": "m3",
        r"^m\^3$": "m3",
        r"^m³.*$": "m3",
        
        # Carbon dioxide equivalents with various notations
        r"^kg.*co.*?2.*(?:eq|äq|aeq|äqv|eqv|[- ]?äquiv|\s?equivalent).*$": "kg CO2 eq",
        r"^kg.*co.*?\(?2\)?.*(?:eq|äq|aeq|äqv|eqv|[- ]?äquiv|\s?equivalent).*$": "kg CO2 eq",
        r"^kg.*co_?\(?2\)?[- ]?äq\.?$": "kg CO2 eq",
        r"^kg\s*co_?\(?2\)?(?:\s|-|_).*$": "kg CO2 eq",
        
        # CFC equivalents
        r"^kg.*(?:cfc|r)\s*11.*(?:eq|äq|aeq|äqv|eqv|[- ]?äquiv|\s?equivalent).*$": "kg CFC-11 eq",
        r"^kg.*(?:cfc|r)11.*(?:eq|äq|aeq|äqv|eqv|[- ]?äquiv|\s?equivalent).*$": "kg CFC-11 eq",
        r"^kg\s*cfc-11\s*eq\.?$": "kg CFC-11 eq",
        
        # Phosphorus equivalents
        r"^kg.*p(?:[ -]?eq|\s?äq|\s?aeq|\s?äqv|\s?eqv|[- ]?äquiv|\s?equivalent).*$": "kg P eq",
        r"^kg.*p[- ]?äq.*$": "kg P eq",
        r"^kg.*phosphat.*$": "kg PO4 eq",
        
        # NMVOC equivalents
        r"^kg.*n.*mvoc.*(?:eq|äq|aeq|äqv|eqv|[- ]?äquiv|\s?equivalent).*$": "kg NMVOC eq",
        r"^kg.*nmvoc.*$": "kg NMVOC eq",
        
        # Ethene/Ethylene equivalents
        r"^kg.*(?:ethen|ethene|ethylen).*(?:eq|äq|aeq|äqv|eqv|[- ]?äquiv|\s?equivalent).*$": "kg C2H2 eq",
        r"^kg.*(?:ethen|ethene|ethylen)[- ]?äq.*$": "kg C2H2 eq",
        r"^kg.*c2h[24].*(?:eq|äq|aeq|äqv|eqv|[- ]?äquiv|\s?equivalent).*$": "kg C2H2 eq",
        
        # Antimony equivalents
        r"^kg.*sb.*(?:eq|äq|aeq|äqv|eqv|[- ]?äquiv|\s?equivalent).*$": "kg Sb eq",
        
        # Sulfur dioxide equivalents
        r"^kg.*so.*?2.*(?:eq|äq|aeq|äqv|eqv|[- ]?äquiv|\s?equivalent).*$": "kg SO2 eq",
        r"^kg.*so_?\(?2\)?.*(?:eq|äq|aeq|äqv|eqv|[- ]?äquiv|\s?equivalent).*$": "kg SO2 eq",
        
        # Phosphate equivalents
        r"^kg.*po.*?4.*(?:eq|äq|aeq|äqv|eqv|[- ]?äquiv|\s?equivalent).*$": "kg PO4 eq",
        r"^kg.*po_?\(?4\)?.*(?:eq|äq|aeq|äqv|eqv|[- ]?äquiv|\s?equivalent).*$": "kg PO4 eq",
        r"^kg.*po.*\(?3-?\)?.*(?:eq|äq|aeq|äqv|eqv|[- ]?äquiv|\s?equivalent).*$": "kg PO4 eq",
        r"^kg.*\(po4\)3-.*$": "kg PO4 eq",
        r"^kg.*phosphate.*$": "kg PO4 eq",
        
        # Nitrogen equivalents
        r"^kg.*n.*(?:eq|äq|aeq|äqv|eqv|[- ]?äquiv|\s?equivalent).*$": "kg N eq",
        r"^mol.*n.*(?:eq|äq|aeq|äqv|eqv|[- ]?äquiv|\s?equivalent).*$": "mol N eq",
        r"^mole.*n.*(?:eq|äq|aeq|äqv|eqv|[- ]?äquiv|\s?equivalent).*$": "mol N eq",
        
        # Hydrogen ion equivalents
        r"^mol.*h.*[\+\^].*(?:eq|äq|aeq|äqv|eqv|[- ]?äquiv|\s?equivalent).*$": "mol H+ eq",
        r"^mole.*h.*[\+\^].*(?:eq|äq|aeq|äqv|eqv|[- ]?äquiv|\s?equivalent).*$": "mol H+ eq",
        r"^mol.*h.*[-\+](?:eq|äq|aeq|äqv|eqv|[- ]?äquiv|\s?equivalent).*$": "mol H+ eq",
        
        # Uranium equivalents
        r"^k?bq.*u235.*(?:eq|äq|aeq|äqv|eqv|[- ]?äquiv|\s?equivalent).*$": "kBq U235 eq",
        
        # World water equivalents
        r"^m.*?3.*world.*(?:eq|äq|aeq|äqv|eqv|[- ]?äquiv|\s?equivalent).*(?:deprived|entzogen)?$": "m3 world eq deprived",
        r"^m\^?\(?3\)?.*w(?:orld|elt).*(?:eq|äq|aeq|äqv|eqv|[- ]?äquiv|\s?equivalent).*(?:deprived|entzogen)$": "m3 world eq deprived",
        
        # Disease incidence
        r"^disease\s*incidence$": "disease incidence",
        r"^krankheitsfälle$": "disease incidence",
        
        # Other specific units
        r"^ctuh$": "CTUh",
        r"^ctue$": "CTUe",
        r"^sqp$": "SQP",
        r"^dimensionless$": "dimensionless",
        
        # Simple units
        r"^-?$": "-",
        r"^mj$": "MJ",
        r"^kg$": "kg",
        
        # Per unit conversions
        r"^kg\/pce$": "kg/pce",
        r"^kg\s*\/\s*pce$": "kg/pce",
        
        # Compound units with divisions or per - volume-based
        r"^kg\s*\/\s*m[\^]?3$": "kg/m3",
        r"^kg\s*\/\s*m3$": "kg/m3",
        r"^kg\s*per\s*m3$": "kg/m3",
        r"^kg\s*per\s*m[\^]?3$": "kg/m3",
        
        # Compound units with divisions or per - area-based
        r"^kg\s*\/\s*m[\^]?2$": "kg/m2",
        r"^kg\s*\/\s*m2$": "kg/m2",
        r"^kg\s*per\s*m2$": "kg/m2",
        r"^kg\s*per\s*m[\^]?2$": "kg/m2"
    }

    for pattern, standard in mappings.items():
        if re.match(pattern, unit):
            return standard
    
    if unit not in not_found_units:
        print(f"Unit '{unit}' not found in standard mappings.")
        not_found_units.append(unit)
    return unit

def connect_to_db():
    """Connect to the PostgreSQL database and return the connection."""
    conn = psycopg2.connect(
        host=os.getenv("DB_HOST", "localhost"),
        database=os.getenv("DB_NAME"),
        user=os.getenv("DB_USER"),
        password=os.getenv("DB_PASSWORD"),
        port=os.getenv("DB_PORT", "5432")
    )
    return conn

def get_products(conn):
    """Get products with process_id, name_en, and category levels."""
    cursor = conn.cursor()
    cursor.execute("""
        SELECT 
            process_id, 
            name_en, 
            category_level_1, 
            category_level_2, 
            category_level_3
        FROM 
            products
    """)
    products = cursor.fetchall()
    cursor.close()
    
    # Convert to DataFrame
    products_df = pd.DataFrame(products, columns=[
        'process_id', 'name_en', 'category_level_1', 
        'category_level_2', 'category_level_3'
    ])
    
    return products_df

def get_lcia_data(conn):
    """Get LCIA results and module amounts."""
    cursor = conn.cursor()
    cursor.execute("""
        SELECT 
            lr.process_id,
            lr.indicator_key,
            lr.unit,
            lma.module,
            lma.scenario,
            lma.amount
        FROM 
            lcia_results lr
        JOIN 
            lcia_moduleamounts lma ON lr.lcia_id = lma.lcia_id
    """)
    lcia_data = cursor.fetchall()
    cursor.close()
    
    # Convert to DataFrame
    lcia_df = pd.DataFrame(lcia_data, columns=[
        'process_id', 'indicator_key', 'unit', 
        'module', 'scenario', 'amount'
    ])

    # Normalize unit values
    lcia_df['unit'] = lcia_df['unit'].apply(normalize_unit)

    # Add type column to identify as LCIA
    lcia_df['type'] = 'lcia'
    
    return lcia_df

def get_exchange_data(conn):
    """Get exchanges and module amounts."""
    cursor = conn.cursor()
    cursor.execute("""
        SELECT 
            e.process_id,
            e.indicator_key,
            e.unit,
            ema.module,
            ema.scenario,
            ema.amount
        FROM 
            exchanges e
        JOIN 
            exchange_moduleamounts ema ON e.exchange_id = ema.exchange_id
    """)
    exchange_data = cursor.fetchall()
    cursor.close()
    
    # Convert to DataFrame
    exchange_df = pd.DataFrame(exchange_data, columns=[
        'process_id', 'indicator_key', 'unit', 
        'module', 'scenario', 'amount'
    ])

    # Normalize unit values
    exchange_df['unit'] = exchange_df['unit'].apply(normalize_unit)

    # Add type column to identify as exchange
    exchange_df['type'] = 'exchange'
    
    return exchange_df

def get_material_properties(conn):
    """Get material properties and prepare them for normalization with indexed naming."""
    cursor = conn.cursor()
    cursor.execute("""
        SELECT 
            process_id,
            property_name,
            value,
            units,
            description
        FROM 
            material_properties
    """)
    material_prop_data = cursor.fetchall()
    cursor.close()
    
    # Convert to DataFrame
    material_prop_df = pd.DataFrame(material_prop_data, columns=[
        'process_id', 'property_name', 'value', 'units', 'description'
    ])
    
    # Convert value column to numeric if possible
    material_prop_df['value'] = pd.to_numeric(material_prop_df['value'], errors='coerce')
    
    # Normalize units values
    material_prop_df['units'] = material_prop_df['units'].apply(normalize_unit)
    
    # Create indexed material property columns
    processed_material_props = {}
    
    # Group by process_id
    for process_id, group in material_prop_df.groupby('process_id'):
        prop_count = 1
        
        # For each material property in the process
        for _, row in group.iterrows():
            material_prop_prefix = f"material_prop_{prop_count}"
            
            # Initialize dict for this process_id if needed
            if process_id not in processed_material_props:
                processed_material_props[process_id] = {}
            
            # Store all material property attributes with indexed naming
            processed_material_props[process_id][f"property_name_{material_prop_prefix}"] = row['property_name']
            processed_material_props[process_id][f"value_{material_prop_prefix}"] = row['value']
            processed_material_props[process_id][f"units_{material_prop_prefix}"] = row['units']
            processed_material_props[process_id][f"description_{material_prop_prefix}"] = row['description']
            
            prop_count += 1
    
    # Convert dictionary to DataFrame
    material_prop_result_df = pd.DataFrame.from_dict(processed_material_props, orient='index')
    material_prop_result_df.reset_index(inplace=True)
    material_prop_result_df.rename(columns={'index': 'process_id'}, inplace=True)
    
    # Also create a DataFrame with the original material properties info for reference
    material_prop_df['material_prop_info'] = material_prop_df.apply(
        lambda row: f"{row['property_name']}: {row['value']} {row['units']} - {row['description']}",
        axis=1
    )
    
    material_prop_info = material_prop_df.groupby('process_id')['material_prop_info'].apply(
        lambda x: '; '.join(x)
    ).reset_index()
    
    return material_prop_result_df, material_prop_info

def get_flow_properties(conn):
    """Get flow properties and identify the reference property for each process."""
    cursor = conn.cursor()
    cursor.execute("""
        SELECT 
            process_id,
            name_en,
            name_de,
            meanamount,
            unit,
            is_reference
        FROM 
            flow_properties
        WHERE
            is_reference = true
    """)
    reference_flow_props = cursor.fetchall()
    cursor.close()
    
    # Convert to DataFrame
    reference_props_df = pd.DataFrame(reference_flow_props, columns=[
        'process_id', 'name_en', 'name_de', 'meanamount', 'unit', 'is_reference'
    ])
    
    # Convert meanamount column to numeric if possible
    reference_props_df['meanamount'] = pd.to_numeric(reference_props_df['meanamount'], errors='coerce')
    
    # Normalize unit values
    reference_props_df['unit'] = reference_props_df['unit'].apply(normalize_unit)
    
    # Rename columns to make their purpose clearer
    reference_props_df = reference_props_df.rename(columns={
        'name_en': 'reference_prop_name_en',
        'name_de': 'reference_prop_name_de',
        'meanamount': 'reference_prop_value',
        'unit': 'reference_prop_unit'
    })
    
    # Keep only needed columns
    reference_props_df = reference_props_df[[
        'process_id', 'reference_prop_name_en', 'reference_prop_name_de', 
        'reference_prop_value', 'reference_prop_unit'
    ]]
    
    return reference_props_df

def calculate_statistics(indicator_data, level=3):
    """
    Calculate statistics (mean, min, max, range) for indicator data grouped by category and module.
    
    Parameters:
    - indicator_data: DataFrame containing the data to calculate statistics from
    - level: Category level to group by (1, 2, or 3)
    
    Returns a DataFrame with statistics for each category and module combination.
    """
    # Determine which category levels to include based on the specified level
    group_cols = ['module']
    if level >= 1:
        group_cols.insert(0, 'category_level_1')
    if level >= 2:
        group_cols.insert(1, 'category_level_2')
    if level >= 3:
        group_cols.insert(2, 'category_level_3')
    
    # Create a grouped summary
    stats = indicator_data.groupby(group_cols)['amount'].agg([
        ('mean', 'mean'),
        ('min', 'min'),
        ('max', 'max')
    ]).reset_index()
    
    # Calculate range
    stats['range'] = stats['max'] - stats['min']
    
    # Sort by category
    sort_cols = [col for col in ['category_level_1', 'category_level_2', 'category_level_3', 'module'] if col in stats.columns]
    stats = stats.sort_values(by=sort_cols)
    
    # Add a column to indicate the aggregation level
    stats['aggregation_level'] = level
    
    return stats

def add_statistics_to_excel(writer, sheet_name, indicator_data, unit=None):
    """
    Add statistics to Excel sheet after product data.
    
    Parameters:
    - writer: ExcelWriter object
    - sheet_name: Name of the sheet to add statistics to
    - indicator_data: DataFrame containing the data to calculate statistics from
    - unit: Optional unit string to display in the statistics section
    """
    # Get the worksheet
    if sheet_name in writer.sheets:
        worksheet = writer.sheets[sheet_name]
    else:
        return  # Sheet doesn't exist, can't add statistics
    
    # Calculate statistics
    stats = calculate_statistics(indicator_data)
    
    if stats.empty:
        return  # No statistics to add
    
    # Find the last row with data
    last_row = worksheet.max_row + 3  # Add 3 empty rows before statistics
    
    # Create styles for the statistics section
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Add statistics section header
    title = "SUMMARY STATISTICS BY CATEGORY AND MODULE"
    if unit:
        title += f" (Unit: {unit})"
    
    worksheet.cell(row=last_row, column=1, value=title)
    header_cell = worksheet.cell(row=last_row, column=1)
    header_cell.font = header_font
    header_cell.fill = header_fill
    
    # Add statistics headers
    headers = ['Category Level 1', 'Category Level 2', 'Category Level 3', 'Module', 'Mean', 'Min', 'Max', 'Range']
    for col_idx, header in enumerate(headers, 1):
        cell = worksheet.cell(row=last_row + 1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    # Add statistics data
    for row_idx, row in enumerate(stats.itertuples(), 1):
        worksheet.cell(row=last_row + 1 + row_idx, column=1, value=row.category_level_1)
        worksheet.cell(row=last_row + 1 + row_idx, column=2, value=row.category_level_2)
        worksheet.cell(row=last_row + 1 + row_idx, column=3, value=row.category_level_3)
        worksheet.cell(row=last_row + 1 + row_idx, column=4, value=row.module)
        worksheet.cell(row=last_row + 1 + row_idx, column=5, value=row.mean)
        worksheet.cell(row=last_row + 1 + row_idx, column=6, value=row.min)
        worksheet.cell(row=last_row + 1 + row_idx, column=7, value=row.max)
        worksheet.cell(row=last_row + 1 + row_idx, column=8, value=row.range)
        
        # Apply borders to all cells in the row
        for col_idx in range(1, 9):
            worksheet.cell(row=last_row + 1 + row_idx, column=col_idx).border = thin_border

def create_all_statistics_sheet(writer, data_by_indicator):
    """
    Create a sheet with all statistics for all indicators at different category levels.
    
    Parameters:
    - writer: ExcelWriter object
    - data_by_indicator: Dictionary with indicator keys and their corresponding DataFrames
    """
    # Create a blank worksheet
    if 'All Statistics' in writer.sheets:
        worksheet = writer.sheets['All Statistics']
    else:
        # Create a new blank DataFrame and write it to initialize the sheet
        pd.DataFrame().to_excel(writer, sheet_name='All Statistics')
        worksheet = writer.sheets['All Statistics']
    
    # Clear any existing content by removing all rows
    for row in worksheet.iter_rows():
        for cell in row:
            cell.value = None
    
    # Start at row 1
    current_row = 1
    
    # Create styles for the statistics section
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    subheader_fill = PatternFill(start_color="E6F2FF", end_color="E6F2FF", fill_type="solid")
    level1_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    level2_fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Add main statistics section header
    worksheet.cell(row=current_row, column=1, value="SUMMARY STATISTICS FOR ALL INDICATORS")
    header_cell = worksheet.cell(row=current_row, column=1)
    header_cell.font = header_font
    header_cell.fill = header_fill
    
    current_row += 2  # Add space after the main header
    
    # Process each indicator
    for indicator, indicator_data in data_by_indicator.items():
        if indicator_data.empty:
            continue
            
        # Add indicator header
        indicator_unit = indicator_data['unit'].iloc[0] if not indicator_data.empty else ""
        
        worksheet.cell(row=current_row, column=1, value=f"INDICATOR: {indicator} ({indicator_unit})")
        indicator_header = worksheet.cell(row=current_row, column=1)
        indicator_header.font = header_font
        indicator_header.fill = subheader_fill
        
        current_row += 1
        
        # Calculate statistics for all three levels
        all_stats = []
        
        for level in [3, 2, 1]:
            level_stats = calculate_statistics(indicator_data, level)
            if not level_stats.empty:
                all_stats.append(level_stats)
        
        if all_stats:
            # Combine all stats
            combined_stats = pd.concat(all_stats, ignore_index=True)
            
            # Add level header based on aggregation_level
            current_row += 1
            worksheet.cell(row=current_row, column=1, value="AGGREGATION LEVEL")
            worksheet.cell(row=current_row, column=1).font = header_font
            worksheet.cell(row=current_row, column=1).fill = header_fill
            
            # Define headers based on the combined stats columns
            headers = []
            if 'category_level_1' in combined_stats.columns:
                headers.append('Category Level 1')
            if 'category_level_2' in combined_stats.columns:
                headers.append('Category Level 2')
            if 'category_level_3' in combined_stats.columns:
                headers.append('Category Level 3')
            
            headers.extend(['Module', 'Mean', 'Min', 'Max', 'Range', 'Aggregation Level'])
            
            # Add headers
            for col_idx, header in enumerate(headers, 1):
                cell = worksheet.cell(row=current_row, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
            
            # Add all statistics data
            for row_idx, row in enumerate(combined_stats.itertuples(), 1):
                # Determine which columns to include based on aggregation level
                col_offset = 0
                
                if hasattr(row, 'category_level_1'):
                    worksheet.cell(row=current_row + row_idx, column=1 + col_offset, value=row.category_level_1)
                    col_offset += 1
                
                if hasattr(row, 'category_level_2'):
                    worksheet.cell(row=current_row + row_idx, column=1 + col_offset, value=row.category_level_2)
                    col_offset += 1
                
                if hasattr(row, 'category_level_3'):
                    worksheet.cell(row=current_row + row_idx, column=1 + col_offset, value=row.category_level_3)
                    col_offset += 1
                
                # Add the rest of the data
                worksheet.cell(row=current_row + row_idx, column=1 + col_offset, value=row.module)
                worksheet.cell(row=current_row + row_idx, column=2 + col_offset, value=row.mean)
                worksheet.cell(row=current_row + row_idx, column=3 + col_offset, value=row.min)
                worksheet.cell(row=current_row + row_idx, column=4 + col_offset, value=row.max)
                worksheet.cell(row=current_row + row_idx, column=5 + col_offset, value=row.range)
                worksheet.cell(row=current_row + row_idx, column=6 + col_offset, value=row.aggregation_level)
                
                # Apply style based on aggregation level
                row_fill = None
                if row.aggregation_level == 1:
                    row_fill = level1_fill
                elif row.aggregation_level == 2:
                    row_fill = level2_fill
                
                # Apply borders and fill to all cells in the row
                for col_idx in range(1, len(headers) + 1):
                    cell = worksheet.cell(row=current_row + row_idx, column=col_idx)
                    cell.border = thin_border
                    if row_fill:
                        cell.fill = row_fill
                
            # Update current_row for the next indicator
            current_row += len(combined_stats) + 3  # +3 for header and spacing
        else:
            # If no statistics, just move to the next indicator
            current_row += 2

def main():
    # Connect to the database
    conn = connect_to_db()
    
    try:
        # Get all required data
        products_df = get_products(conn)
        lcia_df = get_lcia_data(conn)
        exchange_df = get_exchange_data(conn)
        
        # Get material properties
        material_prop_df, material_prop_info = get_material_properties(conn)
        
        # Get reference flow properties
        reference_props_df = get_flow_properties(conn)
        
        # Combine LCIA and exchange data
        combined_data = pd.concat([lcia_df, exchange_df])
        
        # Merge with products data
        result = pd.merge(
            combined_data, 
            products_df,
            on='process_id',
            how='left'
        )
        
        # Merge with reference flow properties
        result = pd.merge(
            result,
            reference_props_df,
            on='process_id',
            how='left'
        )
        
        # Create normalized columns based on reference properties
        result['normalized_amount_reference_prop'] = result.apply(
            lambda row: row['amount'] / row['reference_prop_value'] 
            if pd.notnull(row.get('reference_prop_value')) and row.get('reference_prop_value') != 0 
            else None, 
            axis=1
        )
        
        # Create normalized unit combining indicator unit with reference property unit
        result['normalized_unit_reference_prop'] = result.apply(
            lambda row: f"{row['unit']} / {row.get('reference_prop_unit', '')}" 
            if pd.notnull(row.get('reference_prop_unit')) 
            else None,
            axis=1
        )
        
        # Merge with material properties data
        result = pd.merge(
            result,
            material_prop_df,
            on='process_id',
            how='left'
        )
            
        # Create normalized amount columns for each material property
        material_prop_indices = set()
            
        # Identify all material property indices (material_prop_1, material_prop_2, etc.)
        for col in result.columns:
            if col.startswith('value_material_prop_'):
                material_prop_idx = col.split('value_')[1]  # Gets "material_prop_X"
                material_prop_indices.add(material_prop_idx)
            
        # Sort the material property indices to ensure consistent ordering
        material_prop_indices = sorted(list(material_prop_indices))
            
        # Create normalized columns for each material property
        for material_prop_idx in material_prop_indices:
            value_col = f"value_{material_prop_idx}"
            if value_col in result.columns:
                # Create normalized amount column
                result[f"normalized_amount_{material_prop_idx}"] = result.apply(
                    lambda row: row['amount'] / row[value_col] 
                    if pd.notnull(row.get(value_col)) and row.get(value_col) != 0 
                    else None, 
                    axis=1
                )
                
                # Create normalized unit column
                result[f"normalized_unit_{material_prop_idx}"] = result.apply(
                    lambda row: f"{row['unit']} / {row.get(f'units_{material_prop_idx}', '')}"
                    if pd.notnull(row.get(f'units_{material_prop_idx}')) 
                    else None,
                    axis=1
                )
        
        # Define the columns we want to include in each output
        output_columns = [
            'process_id', 'name_en', 'category_level_1', 'category_level_2', 'category_level_3',
            'module', 'scenario', 'amount', 'unit',
            'reference_prop_name_en', 'reference_prop_value', 'reference_prop_unit',
            'normalized_amount_reference_prop', 'normalized_unit_reference_prop'
        ]
        
        # Add material property columns if they exist
        for material_prop_idx in material_prop_indices:
            if f"normalized_amount_{material_prop_idx}" in result.columns:
                output_columns.extend([
                    f"property_name_{material_prop_idx}",
                    f"value_{material_prop_idx}",
                    f"units_{material_prop_idx}",
                    f"normalized_amount_{material_prop_idx}",
                    f"normalized_unit_{material_prop_idx}"
                ])
        
        # Filter output columns to only include those that exist
        output_columns = [col for col in output_columns if col in result.columns]
        
        # Get unique indicators
        unique_indicators = sorted(result['indicator_key'].dropna().unique())
        
        print(f"Creating Excel file with {len(unique_indicators)} indicator tabs...")
        
        # Create a new Excel workbook
        excel_filename = "indicators_by_category_with_stats.xlsx"
        
        # Dictionary to store data by indicator for summary stats
        indicator_data_dict = {}
        
        # Use pandas ExcelWriter with openpyxl engine
        with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
            # First process all indicators
            for indicator in unique_indicators:
                if pd.isna(indicator) or indicator == "":
                    continue
                    
                # Filter data for this indicator
                indicator_data = result[result['indicator_key'] == indicator].copy()
                
                # Store indicator data for later use in summary stats
                indicator_data_dict[indicator] = indicator_data
                
                # If no data for this indicator, skip
                if indicator_data.empty:
                    continue
                
                print(f"  - Processing indicator: {indicator}")
                
                # Sort by category and name
                indicator_data = indicator_data.sort_values(
                    by=['category_level_1', 'category_level_2', 'category_level_3', 'name_en']
                )
                
                # Create a valid sheet name (max 31 chars, no special chars)
                sheet_name = re.sub(r'[^\w\-_ ]', '', indicator)[:31]
                
                # Instead of directly writing the DataFrame, we'll add empty rows between categories
                df_with_gaps = []
                last_category = None
                
                # Add header row
                df_with_gaps.append(pd.Series({col: col for col in output_columns}))
                
                # Process rows with category separators
                for _, row in indicator_data.iterrows():
                    current_category = (
                        row['category_level_1'], 
                        row['category_level_2'], 
                        row['category_level_3']
                    )
                    
                    # If this is a new category, add an empty row
                    if last_category is not None and current_category != last_category:
                        # Add empty row (all values None)
                        df_with_gaps.append(pd.Series({col: None for col in output_columns}))
                    
                    # Add the data row
                    df_with_gaps.append(pd.Series({col: row.get(col) for col in output_columns}))
                    
                    # Update last category
                    last_category = current_category
                
                # Convert list of Series to DataFrame
                df_to_write = pd.DataFrame(df_with_gaps)
                
                # Write to Excel sheet
                df_to_write.to_excel(writer, sheet_name=sheet_name, header=False, index=False)
                
                # Get the unit for this indicator (assuming all rows have the same unit)
                indicator_unit = indicator_data['unit'].iloc[0] if not indicator_data.empty else ""
                
                # Add statistics to the sheet
                add_statistics_to_excel(writer, sheet_name, indicator_data, indicator_unit)
                
            # Create the 'All Statistics' tab instead of 'All Data' and 'Summary Statistics'
            print("  - Creating 'All Statistics' tab...")
            
            # Create a comprehensive statistics sheet with all levels of aggregation
            create_all_statistics_sheet(writer, indicator_data_dict)
            
        print(f"Excel file created successfully: {excel_filename}")
        
    except Exception as e:
        print(f"An error occurred: {e}")
        # For better debugging, print the full traceback
        traceback.print_exc()
    
    finally:
        conn.close()

if __name__ == "__main__":
    main()